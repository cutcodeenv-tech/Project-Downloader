#!/usr/bin/env python3
"""
Скрипт для парсинга ссылок из Excel и добавления прямых ссылок на изображения.
Принимает Excel файл, анализирует ссылки в ячейках, парсит сайты,
находит все прямые ссылки на картинки и добавляет их в те же ячейки.
"""

import re
import sys
from urllib.parse import urljoin, urlparse
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter


def is_image_url(url):
    """Проверяет, является ли URL ссылкой на изображение."""
    image_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.svg', '.ico')
    parsed = urlparse(url)
    path = parsed.path.lower()
    return any(path.endswith(ext) for ext in image_extensions)


def is_content_image(url):
    """
    Проверяет, является ли URL ссылкой на реальное фото/контент, а не элемент интерфейса.
    Отсеивает иконки, логотипы, UI-элементы, маленькие картинки.
    """
    if not url:
        return False

    url_lower = url.lower()

    # Ключевые слова, указывающие на UI элементы (иконки, логотипы, кнопки и т.д.)
    ui_keywords = [
        'icon', 'logo', 'avatar', 'user', 'profile', 'thumb', 'thumbnail',
        'sprite', 'spinner', 'loader', 'bg-', 'background', 'pattern',
        'blank', 'spacer', 'pixel', 'ad', 'banner', 'promo', 'social',
        'share', 'like', 'tweet', 'facebook', 'twitter', 'instagram',
        'youtube', 'linkedin', 'pinterest', 'whatsapp', 'telegram',
        'arrow', 'chevron', 'close', 'menu', 'hamburger', 'search',
        'cart', 'basket', 'account', 'login', 'logout',
        'settings', 'gear', 'cog', 'home', 'house', 'bell', 'notification',
        'calendar', 'date', 'time', 'clock', 'map', 'marker', 'pin',
        'phone', 'mail', 'email', 'envelope', 'print', 'download',
        'upload', 'edit', 'delete', 'trash', 'plus', 'minus', 'check',
        'cross', 'x-mark', 'star', 'heart', 'bookmark', 'flag', 'info',
        'warning', 'error', 'success', 'loading'
    ]

    # Если в URL есть ключевые слова UI - отсеиваем
    if any(keyword in url_lower for keyword in ui_keywords):
        return False

    # Отсев маленьких размеров (если указаны в URL как 50x50, 16x16 и т.д.)
    size_pattern = re.search(r'(\d{1,3})[xX_](\d{1,3})', url)
    if size_pattern:
        w, h = int(size_pattern.group(1)), int(size_pattern.group(2))
        # Считаем фото контентом, если хотя бы одна сторона > 150px
        if w < 150 and h < 150:
            return False

    # SVG часто используются для иконок - отсеиваем их, кроме случаев явного контента
    if '.svg' in url_lower:
        if not any(word in url_lower for word in ['illustration', 'diagram', 'chart', 'photo', 'image', 'cover', 'hero', 'main']):
             return False

    return True


def extract_image_urls_from_page(page_url):
    """Извлекает все прямые ссылки на изображения со страницы, фильтруя UI элементы."""
    image_urls = set()

    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(page_url, headers=headers, timeout=10, verify=True)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')

        # Ищем теги img
        for img_tag in soup.find_all('img'):
            src = img_tag.get('src') or img_tag.get('data-src') or img_tag.get('data-lazy-src')
            if src:
                full_url = urljoin(page_url, src)
                if is_image_url(full_url) and is_content_image(full_url):
                    image_urls.add(full_url)

        # Ищем ссылки в CSS (background-image и т.д.)
        for style_tag in soup.find_all('style'):
            if style_tag.string:
                css_urls = re.findall(r'url\([\'"]?([^\'")]+)[\'"]?\)', style_tag.string)
                for url in css_urls:
                    full_url = urljoin(page_url, url)
                    if is_image_url(full_url) and is_content_image(full_url):
                        image_urls.add(full_url)

        # Ищем в inline стилях
        for tag in soup.find_all(style=True):
            style = tag.get('style', '')
            css_urls = re.findall(r'url\([\'"]?([^\'")]+)[\'"]?\)', style)
            for url in css_urls:
                full_url = urljoin(page_url, url)
                if is_image_url(full_url) and is_content_image(full_url):
                    image_urls.add(full_url)

    except requests.exceptions.RequestException as e:
        print(f"Ошибка при запросе {page_url}: {e}")
    except Exception as e:
        print(f"Ошибка при парсинге {page_url}: {e}")

    return list(image_urls)


def extract_urls_from_text(text):
    """Извлекает все URL из текста."""
    if not text or not isinstance(text, str):
        return []

    url_pattern = r'https?://[^\s<>"{}|\\^`\[\]]+'
    urls = re.findall(url_pattern, text)
    return urls


def process_excel_file(input_file, output_file=None):
    """Обрабатывает Excel файл: находит ссылки, парсит сайты, добавляет изображения."""
    if output_file is None:
        base_name = input_file.rsplit('.', 1)[0]
        output_file = f"{base_name}_with_images.xlsx"

    print(f"Загрузка файла: {input_file}")
    workbook = openpyxl.load_workbook(input_file)

    total_cells_processed = 0
    total_images_found = 0

    for sheet_name in workbook.sheetnames:
        print(f"\nОбработка листа: {sheet_name}")
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    total_cells_processed += 1
                    urls = extract_urls_from_text(cell.value)

                    if urls:
                        all_images = []
                        for url in urls:
                            print(f"  Парсинг: {url}")
                            images = extract_image_urls_from_page(url)
                            if images:
                                print(f"    Найдено изображений: {len(images)}")
                                all_images.extend(images)

                        if all_images:
                            total_images_found += len(all_images)
                            # Добавляем изображения к существующему содержимому
                            existing_text = cell.value
                            images_text = "\n".join(all_images)
                            cell.value = f"{existing_text}\n\nНайденные изображения:\n{images_text}"
                            print(f"    Добавлено в ячейку {cell.coordinate}")

    workbook.save(output_file)
    print(f"\n{'='*50}")
    print(f"Обработано ячеек: {total_cells_processed}")
    print(f"Всего найдено изображений: {total_images_found}")
    print(f"Результат сохранен в: {output_file}")
    print(f"{'='*50}")

    return output_file


def main():
    # Запрашиваем путь к файлу у пользователя
    input_file = input("Введите путь к входному Excel файлу: ").strip()

    if not input_file:
        print("Ошибка: Путь к файлу не указан.")
        sys.exit(1)

    output_file = input("Введите путь для сохранения результата (или нажмите Enter для имени по умолчанию): ").strip()

    if not output_file:
        output_file = None

    try:
        process_excel_file(input_file, output_file)
    except FileNotFoundError:
        print(f"Ошибка: Файл '{input_file}' не найден.")
        sys.exit(1)
    except Exception as e:
        print(f"Ошибка: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
